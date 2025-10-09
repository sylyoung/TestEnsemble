# -*- coding: utf-8 -*-
# @Time    : 2025/8/24
# @Author  : Siyang Li and Chenhao Liu
# @File    : regression.py
# implementation for our paper "StackingNet: collective inference across independent AI foundation models"
# this file conducts combination methods for regression datasets of research paper review and CFD

import sys, os, argparse, random, re
import numpy as np
import pandas as pd
import torch
import torch.nn as nn
import torch.optim as optim
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import MinMaxScaler
from sklearn.metrics import mean_absolute_error, mean_squared_error
from openpyxl import load_workbook
from pathlib import Path

def save_predictions_excel(save_dir, dataset_name, task, split,
                           test_ids, y_true, model_names, base_preds,
                           method_preds_dict):
    """
    Save ground-truth, base model predictions, and ensemble predictions.

    Args:
        save_dir : str
        dataset_name : str
        task : str
        split : str (e.g. "supervised" or "unsupervised")
        test_ids : (n,) array of paper IDs (optional, can be None)
        y_true : (n,) array of ground-truth values
        model_names : list of str (base model names)
        base_preds : (n_models, n) ndarray of base predictions
        method_preds_dict : dict, {method_name: (n,) array of predictions}
    """
    os.makedirs(save_dir, exist_ok=True)
    out_path = os.path.join(save_dir, f"{dataset_name}_{task}_{split}_preds.xlsx")

    df = pd.DataFrame({"y_true": y_true})
    if test_ids is not None:
        df.insert(0, "paperID", test_ids)

    # Add base models
    for i, m in enumerate(model_names):
        df[m] = base_preds[i]

    # Add ensemble/stacking predictions
    for name, preds in method_preds_dict.items():
        df[name] = preds

    df.to_excel(out_path, index=False)
    print(f"Saved predictions to {out_path}")


def _to_number(x):
    import math, re
    if x is None or (isinstance(x, float) and math.isnan(x)):
        return None
    if isinstance(x, (int, float)):
        return x
    if isinstance(x, str):
        s = x.strip()
        # accept pure numeric strings
        if re.fullmatch(r"-?\d+(?:\.\d+)?", s):
            return int(s) if re.fullmatch(r"-?\d+", s) else float(s)
    return None


def _extract_rating(val):
    """
    Accepts:
      - numbers (3, 4.5)
      - strings "7", " 9.5 "
      - dict-like strings "{'value': '3: reject...'}" or '{"value":"..."}'
      - other strings containing a number, like "5: marginally below..."
    Returns int/float or None.
    """
    import re, json
    from ast import literal_eval

    # 1) Direct numeric
    num = _to_number(val)
    if num is not None:
        return num

    # 2) Try JSON-ish dicts with 'value'/'rating'/'score'
    if isinstance(val, str):
        s = val.strip()
        for parser in (literal_eval, lambda t: json.loads(re.sub(r"'", '"', t))):
            try:
                obj = parser(s)
                if isinstance(obj, dict):
                    for k in ("value", "rating", "score"):
                        if k in obj:
                            inner = obj[k]
                            # try strict number first
                            num = _to_number(inner)
                            if num is not None:
                                return num
                            # then extract leading number from "3: reject ..." etc.
                            if isinstance(inner, str):
                                m = re.search(r"-?\d+(?:\.\d+)?", inner)
                                if m:
                                    ns = m.group(0)
                                    return int(ns) if re.fullmatch(r"-?\d+", ns) else float(ns)
                break
            except Exception:
                continue

        # 3) Fallback: extract first number from raw string
        m = re.search(r"-?\d+(?:\.\d+)?", s)
        if m:
            ns = m.group(0)
            return int(ns) if re.fullmatch(r"-?\d+", ns) else float(ns)

    return None


def build_id_to_ratings(ids_path,
                        reviews_path,
                        id_col="paperID",
                        rating_col="rating",
                        sheet_ids=0,
                        sheet_reviews=0):
    '''
    ids_path: Excel path containing a column `id_col`
    reviews_path: Excel path containing columns `id_col`, `rating_col`
    sheet_*: sheet name or index if needed (default first sheet)
    '''
    # Read Excel
    ids_df = pd.read_excel(ids_path, sheet_name=sheet_ids, dtype={id_col: str})
    rev_df = pd.read_excel(reviews_path, sheet_name=sheet_reviews, dtype={id_col: str})

    # Basic checks
    if id_col not in ids_df.columns:
        raise ValueError(f"'{id_col}' column missing in {ids_path}")
    for col in (id_col, rating_col):
        if col not in rev_df.columns:
            raise ValueError(f"'{col}' column missing in {reviews_path}")

    # Normalize IDs
    ids_df[id_col] = ids_df[id_col].astype(str).str.strip()
    rev_df[id_col] = rev_df[id_col].astype(str).str.strip()

    # Clean ratings to numeric (where possible)
    ratings_clean = rev_df[rating_col].apply(_extract_rating)

    # Keep only rows with ID we care about and non-null rating
    want_ids = set(ids_df[id_col].dropna().unique())
    mask = rev_df[id_col].isin(want_ids) & ratings_clean.notna()
    rev_subset = rev_df.loc[mask, [id_col]].copy()
    rev_subset[rating_col] = ratings_clean[mask]

    # Group to lists
    grouped = rev_subset.groupby(id_col)[rating_col].apply(list)

    # Ensure all requested IDs exist in output
    mapping = {pid: grouped.get(pid, []) for pid in ids_df[id_col].dropna().unique()}

    # remove the final row with MAE
    mapping.pop("MAE")

    return mapping


def seed_everything(args):
    random.seed(args.seed)
    torch.manual_seed(args.seed)
    torch.cuda.manual_seed_all(args.seed)
    np.random.seed(args.seed)
    os.environ['PYTHONHASHSEED'] = str(args.seed)


class StackingRegression(nn.Module):
    def __init__(self, num_models: int):
        super().__init__()
        # uniform initialization: scale 1, bias 0
        self.scales = nn.Parameter(torch.ones(num_models) / num_models)
        self.biases = nn.Parameter(torch.zeros(num_models))

    def forward(self, x):
        scales = torch.clamp(self.scales, min=0.0)
        biases = torch.clamp(self.biases, min=0.0)
        return torch.sum(scales * (x + biases), dim=1)

def train_stacking_regression(pred_train, labels_train, pred_test, args, device=None):
    num_samples, num_models = pred_train.shape

    model = StackingRegression(num_models).to(device)
    optimizer = optim.Adam(model.parameters(), lr=args.lr)
    criterion = nn.MSELoss()

    X_train_t = torch.tensor(pred_train, dtype=torch.float32)
    y_train_t = torch.tensor(labels_train, dtype=torch.float32)
    X_test_t = torch.tensor(pred_test, dtype=torch.float32)
    # y_test_t = torch.tensor(y_test, dtype=torch.float32)

    if device is not None:
        X_train_t, y_train_t = X_train_t.to(device), y_train_t.to(device)
        X_test_t = X_test_t.to(device)

    model.train()
    for _ in range(args.epoch):
        preds = model(X_train_t)
        loss = criterion(preds, y_train_t)
        optimizer.zero_grad()
        loss.backward()
        optimizer.step()

    print('Supervised Relu Training Stacking Model Weights:')
    np.set_printoptions(precision=3, suppress=True)
    print('model.scales:', model.scales.detach())
    print('model.biases:', model.biases.detach())

    with torch.no_grad():
        preds = model(X_test_t)
    return preds.to('cpu').detach().numpy()


def save_results(raw_df, save_path, args):
    # First write DataFrame (with mean±std strings) to Excel
    with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
        raw_df.to_excel(writer, sheet_name='Raw Absolute Error')

    # if args.dataset_name == 'CFD':
    #     labels = labels_test.reshape(1, -1)
    #     Relu_preds = Relu_preds.reshape(1, -1)
    #     combined_pred = np.vstack((labels_test, pred_test, Relu_preds))
    #     combined_df = pd.DataFrame(combined_pred.T,
    #                                columns=['label'] + model_names + ['stacking'])
    #     combined_df = pd.concat([gender_ethnicity.reset_index(drop=True), combined_df], axis=1)
    #     combined_df.to_excel(f'/mnt/data3/lch/Golden_task/regression/results/CFD/{task}.xlsx', index=False)

    # Open workbook for post-processing
    wb = load_workbook(save_path)
    ws = wb['Raw Absolute Error']

    # Write header for Dataset Mean column (P1, but safer to use max_column+1)
    mean_col = ws.max_column + 1
    ws.cell(row=1, column=mean_col).value = "Dataset Mean"

    # For each method row, compute average of numeric MEANS across tasks
    for i in range(2, ws.max_row + 1):
        row_values = []
        for j in range(2, mean_col):  # skip index col, stop before new col
            val = ws.cell(row=i, column=j).value
            if isinstance(val, str):
                match = re.match(r"([\d\.]+)", val)  # capture leading mean number
                if match:
                    row_values.append(float(match.group(1)))
        if row_values:
            ws.cell(row=i, column=mean_col).value = sum(row_values) / len(row_values)

    wb.save(save_path)


def make_unique_index(index):
    unique_index = []
    counts = {}
    for item in index:
        if item in counts:
            counts[item] += 1
            unique_index.append(f"{item}_{counts[item]}")
        else:
            counts[item] = 0
            unique_index.append(item)
    return unique_index


def aggregate_runs(scores_list, row_index, col_name, omit_std_for=None,
                   mean_fmt="{:.3f}", meanstd_fmt="{:.3f}$_{{\\pm{:.3f}}}$",
                   use_sample_std=True):
    """
    scores_list: list of per-run lists/arrays, shape = (num_runs, num_methods)
    row_index: list of method names (len = num_methods)
    col_name: task name (string)
    omit_std_for: set/list of method names for which we show only mean (no ±).
    """
    omit_std_for = set(omit_std_for or [])
    A = np.asarray(scores_list, dtype=float)   # (R, M)

    # Robust to any NaNs (if some runs are missing)
    mean = np.nanmean(A, axis=0)
    ddof = 1 if use_sample_std and A.shape[0] > 1 else 0
    std  = np.nanstd(A, axis=0, ddof=ddof)

    formatted = []
    for m_name, m, s in zip(row_index, mean, std):
        if m_name in omit_std_for:
            formatted.append(mean_fmt.format(m))
        else:
            formatted.append(meanstd_fmt.format(m, s))

    return pd.DataFrame({col_name: formatted}, index=row_index)


def compute_individual_human_mae(test_ids, y_true_test, id_to_ratings):
    abs_errs = []
    missing = []
    for pid, y in zip(test_ids, y_true_test):
        ratings = id_to_ratings.get(str(pid).strip(), [])
        if not ratings:
            missing.append(pid)
            continue
        for r in ratings:
            abs_errs.append(abs(float(r) - float(y)))

    if missing:
        # print a few to see what's up, then hard error
        print(f"Missing human ratings for {len(missing)} test papers. Examples: {list(missing)[:5]}")
        raise ValueError(f"ERROR! No human ratings found for test papers (first: {missing[0]})")

    return float(np.mean(abs_errs))


def run_minmax_supervised(pred_train, pred_test, labels_train, labels_test, scaler, task, args):
    print("#" * 50)
    print(f"[Supervised] Processing task: {task}")

    labels_train = scaler.transform(labels_train.reshape(-1, 1)).flatten()
    labels_test = scaler.transform(labels_test.reshape(-1, 1)).flatten()
    pred_test = np.array([scaler.transform(p.reshape(-1, 1)).flatten() for p in pred_test])
    pred_train = np.array([scaler.transform(p.reshape(-1, 1)).flatten() for p in pred_train])
    labels_test_orig = scaler.inverse_transform(labels_test.reshape(-1, 1))

    avg_pred = np.mean(pred_test, axis=0)
    avg_pred_orig = scaler.inverse_transform(avg_pred.reshape(-1, 1))
    avg_mae = mean_absolute_error(labels_test_orig, avg_pred_orig)

    pred_test = pred_test.T
    pred_train = pred_train.T

    stackingnet_preds = train_stacking_regression(pred_train, labels_train, pred_test, args, device=args.device)
    stackingnet_preds_orig = scaler.inverse_transform(stackingnet_preds.reshape(-1, 1))
    stackingnet_mae = mean_absolute_error(labels_test_orig, stackingnet_preds_orig)

    return avg_mae, stackingnet_mae


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--seed', type=int, default=0)
    parser.add_argument('--dataset_name', type=str, default='research_review', help='dataset name')
    parser.add_argument('--gpu_id', type=int, default=0)
    parser.add_argument('--log', type=str, default='none', help='time log')
    parser.add_argument('--epoch', type=int, default=1000, help='training epochs')
    parser.add_argument('--lr', type=float, default=0.01, help='learning rate')
    parser.add_argument('--unsupervised_weight', type=float, default=0.001, help='weight for loss objective for unlabeled data')
    parser.add_argument('--golden_num', type=float, default=0.5, help='the percentage of the ratio of (labeled data / all data)')
    parser.add_argument('--save_dir', type=str, default="./results/")
    args = parser.parse_args()

    seed_everything(args)
    args.data_dir = './data/' + args.dataset_name + '/'

    if not os.path.isdir(args.data_dir):
        print('assume data located under:', './data/' + args.dataset_name + '/')
        sys.exit(0)

    os.environ["CUDA_VISIBLE_DEVICES"] = '0,1,2,3,4,5,6'
    print(f"Available CUDA devices: {torch.cuda.device_count()}")
    print('torch.cuda.is_available()', torch.cuda.is_available())
    try:
        if torch.cuda.device_count() == 1:
            device = torch.device("cuda:" + str(0) if torch.cuda.is_available() else "cpu")
        else:
            device = torch.device("cuda:" + str(args.gpu_id) if torch.cuda.is_available() else "cpu")
    except:
        device = 'local'
    print('device:', device)
    args.device = device

    if not os.path.isdir(args.save_dir):
        path = Path(args.save_dir)
        path.mkdir(parents=True)

    if args.dataset_name == 'CFD':
        task_arr = ['afraid','angry','babyfaced','feminine','masculine','sad','threatening',
                    'unusual','age','attractive','disgusted','happy','surprised','trustworthy']
    elif args.dataset_name == 'research_review':
        task_arr = ['ICLR2025', 'ICLR2024', 'NeurIPS2024', 'NeurIPS2023']

    all_dfs_sup = []
    all_dfs_unsup = []

    for task in task_arr:

        if args.dataset_name == 'research_review':
            IDS_XLSX = os.path.join(args.data_dir, task + '.xlsx')
            REVIEWS_XLSX = './data/' + 'Paper_Review/' + task + '_reviews.xlsx'
            mapping = build_id_to_ratings(
                IDS_XLSX,
                REVIEWS_XLSX,
                id_col="paperID",
                rating_col="rating",
                sheet_ids=0,
                sheet_reviews=0
            )

        file_path = os.path.join(args.data_dir, task + '.xlsx')
        df = pd.read_excel(file_path)

        if args.dataset_name == 'CFD':
            # Keep Gender/Ethnicity for later (if present)
            gender_ethnicity = None
            if all(c in df.columns for c in ['Gender', 'Ethnicity']):
                gender_ethnicity = df[['Gender', 'Ethnicity']].copy()
            assert gender_ethnicity is not None, print('gender ethnicity not found!')
            df = df.drop(columns=['Gender', 'Ethnicity'])

        if args.dataset_name == 'research_review':
            mae_mask = df['paperID'] == 'MAE'
            df_no_mae = df.loc[~mae_mask].copy()
            paper_ids = df_no_mae['paperID'].to_numpy()
            print('len(paper_ids)', len(paper_ids))

            mae_rows = df[df['paperID'] == 'MAE']
            mae_indices = df[df['paperID'] == 'MAE'].index
            df = df.drop(mae_indices)
            df = df.drop(columns=['paperID'])

        labels = df['label'].values
        predictions_df = df.drop(columns=['label'])
        model_names = list(predictions_df.columns)
        preds = predictions_df.T.values  # (num_models, num_samples)
        print('preds.shape, labels.shape', preds.shape, labels.shape)

        # clip range
        if args.dataset_name == 'research_review':
            min_val = 1.0
            max_val = 10.0
        elif args.dataset_name == 'CFD':
            min_val = 1.0
            max_val = 7.0

        labels_clipped = np.clip(labels, min_val, max_val)
        pred_clipped = np.clip(preds, min_val, max_val)

        # Normalize using fixed min-max range
        # Norm should be executed after split, calculated on training set (though MinMaxNorm does not incur this issue)
        scaler = MinMaxScaler()
        scaler.fit(np.array([[min_val], [max_val]]))

        start_seed = 0

        print('shuffle, and retain 80% data as test data for all random seeds cases')

        preds_T = preds.T  # shape: (num_samples, num_classifiers)

        # split data into 20% retained POOL for labeled "training" data
        # the actual x% labeled data will be further randomly selected in this pool (x < 20)
        # while the 80% fixed test set is used for actual performance evaluation in all cases under all random seeds
        preds_train_T, preds_test_T, labels_train_T, labels_test_T, idx_train, idx_test = train_test_split(
            preds_T, labels, np.arange(len(labels)),
            test_size=0.8,
            random_state=start_seed,
        )

        preds_train, preds_test, labels_train, labels_test = preds_train_T.T, preds_test_T.T, labels_train_T.T, labels_test_T.T

        print('preds_train.shape, labels_train.shape (Pool of Labeled Data):', preds_train.shape,
              labels_train.shape)
        print('preds_test.shape, labels_test.shape:', preds_test.shape, labels_test.shape)

        # --- Normalized [0,1] arrays for later use ---
        labels_test_01 = scaler.transform(labels_test.reshape(-1, 1)).flatten()  # (n_test,)
        preds_test_01 = np.array([scaler.transform(p.reshape(-1, 1)).flatten()
                                  for p in preds_test])  # (num_models, n_test)
        labels_train_01 = scaler.transform(labels_train.reshape(-1, 1)).flatten()  # (n_train,)
        preds_train_01 = np.array([scaler.transform(p.reshape(-1, 1)).flatten()
                                   for p in preds_train])  # (num_models, n_train)

        if args.dataset_name == 'research_review':
            test_ids = paper_ids[idx_test]
            human_individual_mae = compute_individual_human_mae(test_ids, labels_test, mapping)

        # Save the split indices (for pool of labeled data and test data, not for different golden data under random seeds) for reproducibility
        np.savez('./results/' + args.dataset_name + '_' + args.log + '_split_indices.npz', idx_train=idx_train,
                 idx_test=idx_test)

        singles_all = []
        for i, model in enumerate(model_names):
            # compute at full precision (float64)
            score = mean_absolute_error(labels_test.astype(np.float64),
                                        preds_test[i].astype(np.float64))
            # print with formatting ONLY for display
            print(f'Test set ground-truth MAE: {score:.3}  Model_Name: {model}')
            singles_all.append(score)

        # average with full precision, then format for display
        single_avg = float(np.mean(singles_all))
        single_worst = float(np.max(singles_all))
        single_best = float(np.min(singles_all))
        global_model = model_names[int(np.argmin(singles_all))]

        print(f'Test set average (avg over individual base models): {single_avg:.3f}')
        print(f'Test set worst & best single-model MAE: {single_worst:.3f} & {single_best:.3f} '
              f'Best_Model_Name: {global_model}')

        scores_sup = []
        scores_unsup = []

        for seed in range(5):  # for research paper review
            print('#' * 60)
            print('#' * 20, 'random seed:', seed, '#' * 20)
            print('#' * 60)
            args.seed = seed
            seed_everything(args)

            if args.golden_num == 0:
                golden_num = 0
            else:
                golden_num = int(preds.shape[1] // (100 // args.golden_num))
            print('golden_num:', golden_num)

            golden_indices = np.random.RandomState(seed).choice(len(labels_train_T), size=golden_num, replace=False)

            preds_golden = preds_train_T[golden_indices].T
            labels_golden = labels_train_T[golden_indices].T
            print('preds_golden.shape, labels_golden.shape:', preds_golden.shape, labels_golden.shape)
            preds_golden_and_test = np.concatenate([preds_golden, preds_test], axis=1)
            labels_golden_and_test = np.concatenate([labels_golden, labels_test])
            print('preds_golden_and_test.shape, labels_golden_and_test.shape:', preds_golden_and_test.shape,
                  labels_golden_and_test.shape)

            score_sup = run_minmax_supervised(preds_golden, preds_test, labels_golden, labels_test, scaler, task, args)
            score_sup = singles_all + list(score_sup)
            if args.dataset_name == 'research_review':
                score_sup = score_sup + [human_individual_mae]
            scores_sup.append(score_sup)

        method_names_sup = model_names + ['Averaging', 'StackingNet']
        if args.dataset_name == 'research_review':
            method_names_sup += ['Individual-Human']
        df_task_sup = aggregate_runs(
            scores_sup,
            row_index=method_names_sup,
            col_name=task,
            omit_std_for=model_names
        )
        all_dfs_sup.append(df_task_sup)

    final_sup_df = pd.concat(all_dfs_sup, axis=1)
    save_results(final_sup_df, os.path.join(args.save_dir, f"{args.dataset_name}{args.log}_supervised.xlsx"), args)
