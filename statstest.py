"""
We quantify the difference between two models’ class predictions.
Note that this purpose is different from classic statistical tests for machine learning algorithms, which usually operate on evaluation metrics such as accuracy scores.
We use the disagreement rate—the proportion of test samples on which their predicted labels differ.
We then conduct a one-sided proportion test comparing the observed disagreement against a predefined practical threshold δ (the minimum disagreement considered meaningfully large), and reject the null hypothesis at significance level α.
A significant result indicates that the two models’ predictions are sufficiently different.
This is different from CIM paper’s usage of statistical tests in Section IV.E (Statistical Tests), and we believe the current method is more appropriate for our purpose.
"""

import numpy as np

from scipy.stats import norm
from openpyxl import Workbook
from openpyxl.styles import Font

def bh_fdr(pvals):
    """Benjamini–Hochberg FDR adjustment.

    Args:
        pvals: 1D array-like of raw p-values.

    Returns:
        adj_pvals: np.ndarray of BH-adjusted p-values (same shape as input).
    """
    pvals = np.asarray(pvals, dtype=float)
    m = pvals.size
    order = np.argsort(pvals)
    ranked = pvals[order]

    # Compute BH adjusted p-values on the sorted list
    adj = ranked * m / (np.arange(1, m + 1))
    # Enforce monotonicity from largest to smallest
    adj = np.minimum.accumulate(adj[::-1])[::-1]
    adj = np.clip(adj, 0.0, 1.0)

    # Unsort back to original order
    adj_pvals = np.empty_like(adj)
    adj_pvals[order] = adj
    return adj_pvals

def prediction_difference_test(pred1, pred2, delta=0.05, alpha=0.05):
    pred1 = np.asarray(pred1).astype(int)
    pred2 = np.asarray(pred2).astype(int)

    n = len(pred1)
    disagree = np.sum(pred1 != pred2)
    p_hat = disagree / n

    se = np.sqrt(delta * (1 - delta) / n)
    z = (p_hat - delta) / se
    p_value = 1 - norm.cdf(z)

    return {
        "n": n,
        "disagree_rate": p_hat,
        "z": z,
        "p_value": p_value,
    }



names = ['MOSI', 'TUSAS', 'TweetEval', 'WOS']

# LaTeX table row order (SML-OVR vs. ...)
row_order = [
    ('voting', 'Voting'),
    ('wawa', 'WAwA'),
    ('dawidskene', 'Dawid-Skene'),
    ('zhang', 'Zhang et al.'),
    ('glad', 'GLAD'),
    ('mace', 'MACE'),
    ('m-msr', 'M-MSR'),
    ('zencrowd', 'ZenCrowd'),
    ('pm', 'PM'),
    ('la', 'LA'),
    ('laa', 'LAA'),
    ('ebcc', 'EBCC'),
]

# Store BH-adjusted p-values in a table: {row_key: {dataset_name: value}}
all_adj = {k: {} for k, _ in row_order}


delta = 0.05  # threshold on the disagreement rate
alpha = 0.05  # statistical significance level

for data_name in names:
    print("\n" + "#" * 50)
    print("Dataset:", data_name)

    predA = np.loadtxt(f'./results/{data_name}/smlovr.csv')

    # Build compare list in the same order as the LaTeX table rows
    compare_paths = []
    for key, _label in row_order:
        compare_paths.append(f'./results/{data_name}/{key}.csv')

    results = []
    pvals = []

    for path in compare_paths:
        try:
            predB = np.loadtxt(path)
            res = prediction_difference_test(predA, predB, delta=delta, alpha=alpha)
            res["path"] = path
            results.append(res)
            pvals.append(res["p_value"])
        except OSError:
            # Missing baseline file: keep placeholder
            res = {"n": len(predA), "disagree_rate": float('nan'), "z": float('nan'), "p_value": float('nan'), "path": path}
            results.append(res)
            pvals.append(float('nan'))

    # Dataset-wise BH-FDR correction across all baselines in this dataset
    # Handle NaNs by only correcting the finite p-values, then restoring NaNs
    pvals_arr = np.asarray(pvals, dtype=float)
    finite_mask = np.isfinite(pvals_arr)
    adj_pvals = np.full_like(pvals_arr, np.nan, dtype=float)
    if np.any(finite_mask):
        adj_pvals[finite_mask] = bh_fdr(pvals_arr[finite_mask])

    # Print and cache BH-adjusted p-values
    for (key, _label), res, p_adj in zip(row_order, results, adj_pvals):
        res["p_value_adj"] = float(p_adj) if np.isfinite(p_adj) else float('nan')
        all_adj[key][data_name] = res["p_value_adj"]

        print(res["path"])
        print(f"Sample num = {res['n']}")
        print(f"Disagreement rate = {res['disagree_rate']:.4f}" if np.isfinite(res['disagree_rate']) else "Disagreement rate = NaN")
        print(f"Z-value = {res['z']:.4f}" if np.isfinite(res['z']) else "Z-value = NaN")
        print(f"p-value (raw) = {res['p_value']:.6f}" if np.isfinite(res['p_value']) else "p-value (raw) = NaN")
        print(f"p-value (BH)  = {res['p_value_adj']:.6f}" if np.isfinite(res['p_value_adj']) else "p-value (BH)  = NaN")


# -------- Export to Excel (BH-adjusted p-values) --------
wb = Workbook()
ws = wb.active
ws.title = "statstesttext"

# Header row
ws.cell(row=1, column=1, value="SML-OVR vs.").font = Font(bold=True)
for j, data_name in enumerate(names, start=2):
    ws.cell(row=1, column=j, value=data_name).font = Font(bold=True)

# Body rows
for i, (key, label) in enumerate(row_order, start=2):
    ws.cell(row=i, column=1, value=label)
    for j, data_name in enumerate(names, start=2):
        val = all_adj.get(key, {}).get(data_name, np.nan)
        cell = ws.cell(row=i, column=j, value=None if not np.isfinite(val) else float(val))
        cell.number_format = "0.0000"  # keep four decimals
        # Bold if p < 0.05 (as in LaTeX description)
        if np.isfinite(val) and val < 0.05:
            cell.font = Font(bold=True)

# Basic column widths for readability
ws.column_dimensions['A'].width = 18
for col_letter in ['B', 'C', 'D', 'E']:
    ws.column_dimensions[col_letter].width = 12

out_path = "statstesttext_bh_fdr.xlsx"
wb.save(out_path)
print(f"\nSaved BH-FDR table to: {out_path}")
